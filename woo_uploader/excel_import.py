from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook

from .models import IMAGE_EXTENSIONS, WOO_FIELDS, ProductInput, ValidationResult, clean_value, validate_product


@dataclass
class Spreadsheet:
    headers: list[str]
    rows: list[tuple[int, dict[str, str]]]


HEADER_ALIASES = {
    "referencia": "sku",
    "referencia_sku": "sku",
    "titulo": "name",
    "titulo_producto": "name",
    "nombre_producto": "name",
    "precio": "regular_price",
    "precio_regular": "regular_price",
    "precio_normal": "regular_price",
    "precio_rebajado": "sale_price",
    "precio_oferta": "sale_price",
    "descripcion_corta": "short_description",
    "categoria": "categories",
    "categorias": "categories",
    "etiqueta": "tags",
    "etiquetas": "tags",
    "stock": "stock_quantity",
    "cantidad_stock": "stock_quantity",
    "cantidad_de_stock": "stock_quantity",
    "estado_existencias": "stock_status",
    "estado_de_existencias": "stock_status",
}


def infer_mapping(headers: list[str]) -> dict[str, str]:
    """Asocia automáticamente las cabeceras habituales de un Excel a WooCommerce."""
    mapping = {}
    for header in headers:
        normalized = unicodedata.normalize("NFD", header.casefold())
        normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
        mapping[header] = normalized if normalized in WOO_FIELDS else HEADER_ALIASES.get(normalized, "ignore")
    return mapping


def read_spreadsheet(path: str | Path) -> Spreadsheet:
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book.active
    if sheet is None:
        raise ValueError("El Excel no contiene ninguna hoja activa.")
    source_rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean_value(value) for value in next(source_rows)]
    except StopIteration:
        raise ValueError("El Excel está vacío.")
    if not any(headers):
        raise ValueError("La primera fila no contiene cabeceras.")
    if len(set(header for header in headers if header)) != len([header for header in headers if header]):
        raise ValueError("Las cabeceras del Excel deben ser únicas.")
    rows = []
    for row_number, row in enumerate(source_rows, start=2):
        data = {headers[index]: clean_value(value) for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(data.values()):
            rows.append((row_number, data))
    return Spreadsheet(headers, rows)


def find_image(directory: str | Path, image_name: str) -> Path | None:
    if not image_name:
        return None
    directory = Path(directory)
    if not directory.is_dir():
        return None
    needle = image_name.casefold()
    candidates = {needle, Path(needle).stem}
    for path in directory.iterdir():
        if path.is_file() and path.suffix.casefold() in IMAGE_EXTENSIONS:
            if path.name.casefold() in candidates or path.stem.casefold() in candidates:
                return path
    return None


def build_products(
    sheet: Spreadsheet,
    mapping: dict[str, str],
    image_column: str | None = None,
    image_directory: str | Path | None = None,
) -> list[ValidationResult]:
    results = []
    for row_number, row in sheet.rows:
        values = {field: row.get(header, "") for header, field in mapping.items() if field != "ignore"}
        image = find_image(image_directory, row.get(image_column, "")) if image_column and image_directory else None
        results.append(validate_product(ProductInput(values, image, row_number), image_required=False))
    return results

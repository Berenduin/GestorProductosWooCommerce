"""Artefactos de ejecución producidos por los casos de uso."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from platformdirs import user_downloads_path

from .uploads import BatchUploadResult

OUTCOME_NAMES = {"created": "creado", "updated": "actualizado", "skipped": "omitido", "failed": "error"}
PURPLE = "623642"


@dataclass(frozen=True)
class BatchReportPaths:
    csv: Path
    xlsx: Path


def _report_rows(result: BatchUploadResult) -> list[list[str | int]]:
    return [[item.row_number or "", item.sku, OUTCOME_NAMES[item.outcome], item.detail] for item in result.rows]


def write_batch_report(result: BatchUploadResult, directory: Path | None = None) -> Path:
    """Escribe el informe CSV con el nombre histórico de la aplicación."""
    path = (directory or Path.cwd()) / "resultado_subida_lote.csv"
    rows = [["fila", "sku", "resultado", "detalle"], *_report_rows(result)]
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        csv.writer(output).writerows(rows)
    return path


def _write_excel_report(result: BatchUploadResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "Resultado de la subida por lotes"
    summary["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 28
    summary["A3"] = "Estado"
    summary["B3"] = "Cancelado" if result.cancelled else "Proceso terminado"
    counts = result.counts
    for row, (label, key) in enumerate(
        (("Creados", "created"), ("Actualizados", "updated"), ("Omitidos", "skipped"), ("Fallidos", "failed")),
        start=5,
    ):
        summary.cell(row, 1, label)
        summary.cell(row, 2, counts[key])
        summary.cell(row, 1).font = Font(bold=True)
        summary.cell(row, 2).alignment = Alignment(horizontal="right")
    summary["A3"].font = Font(bold=True, color=PURPLE)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 4
    summary.column_dimensions["D"].width = 4

    detail = workbook.create_sheet("Detalle")
    detail.sheet_view.showGridLines = False
    detail.append(["Fila", "SKU", "Resultado", "Detalle"])
    for row in _report_rows(result):
        detail.append([row[0], row[1], str(row[2]).capitalize(), row[3]])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:D{max(detail.max_row, 1)}"
    for cell in detail[1]:
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    outcome_fills = {"Creado": "E2F0D9", "Actualizado": "DDEBF7", "Omitido": "FFF2CC", "Error": "FCE4D6"}
    for row in range(2, detail.max_row + 1):
        result_cell = detail.cell(row, 3)
        result_cell.fill = PatternFill("solid", fgColor=outcome_fills.get(result_cell.value, "FFFFFF"))
        result_cell.font = Font(bold=True)
        detail.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
    detail.column_dimensions["A"].width = 10
    detail.column_dimensions["B"].width = 22
    detail.column_dimensions["C"].width = 16
    detail.column_dimensions["D"].width = 70
    detail.row_dimensions[1].height = 24
    detail.sheet_properties.pageSetUpPr.fitToPage = True
    detail.page_setup.fitToWidth = 1
    detail.page_setup.fitToHeight = 0
    workbook.save(path)


def _excel_destination(directory: Path | None) -> Path:
    destination = directory or Path(user_downloads_path())
    try:
        destination.mkdir(parents=True, exist_ok=True)
        return destination
    except OSError:
        if directory is not None:
            raise
        fallback = Path.cwd()
        fallback.mkdir(parents=True, exist_ok=True)
        return fallback


def write_batch_reports(
    result: BatchUploadResult,
    csv_directory: Path | None = None,
    xlsx_directory: Path | None = None,
    timestamp: datetime | None = None,
) -> BatchReportPaths:
    """Genera el CSV histórico y un informe Excel visible para la persona usuaria."""
    stamp = (timestamp or datetime.now()).strftime("%Y%m%d_%H%M%S")
    excel_directory = _excel_destination(xlsx_directory)
    xlsx_path = excel_directory / f"resultado_subida_lote_{stamp}.xlsx"
    _write_excel_report(result, xlsx_path)
    try:
        csv_path = write_batch_report(result, csv_directory)
    except OSError:
        if csv_directory is not None:
            raise
        csv_path = write_batch_report(result, excel_directory)
    return BatchReportPaths(csv=csv_path, xlsx=xlsx_path)

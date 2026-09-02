from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import pytest

from woo_uploader.models import ProductInput, validate_product
from woo_uploader.services import reports as report_service
from woo_uploader.services.reports import write_batch_report, write_batch_reports
from woo_uploader.services.uploads import BatchUploadService, DuplicateAction, ProductUploadService
from woo_uploader.woocommerce import WooCommerceError


class FakeClient:
    def __init__(self, existing: dict[str, dict] | None = None, fail_skus: set[str] | None = None) -> None:
        self.existing = existing or {}
        self.fail_skus = fail_skus or set()
        self.created: list[dict] = []
        self.updated: list[tuple[int, dict]] = []

    def find_by_sku(self, sku: str):
        return self.existing.get(sku)

    def create_product(self, payload, image_path=None):
        if payload.get("sku", "") in self.fail_skus:
            raise WooCommerceError("fallo remoto")
        self.created.append(payload)
        return {"id": 100 + len(self.created)}

    def update_product(self, product_id, payload, image_path=None):
        self.updated.append((product_id, payload))
        return {"id": product_id}


def valid_product(sku: str, row: int | None = None):
    return validate_product(ProductInput({"name": f"Producto {sku}", "sku": sku, "regular_price": "12,50", "categories": "General", "height": "10"}, row_number=row), required_fields=("name", "sku", "regular_price", "categories", "height"))


def test_product_service_creates_and_resolves_duplicates() -> None:
    client = FakeClient({"OLD": {"id": 8}})
    service = ProductUploadService(lambda: client)
    created = service.upload(valid_product("NEW").product, "draft")
    assert created.outcome == "created"
    assert client.created[0]["type"] == "simple"
    conflict = service.find_conflict(valid_product("OLD").product)
    assert conflict and conflict.existing_product["id"] == 8
    skipped = service.upload(valid_product("OLD").product, "draft", DuplicateAction.SKIP, conflict.existing_product)
    assert skipped.outcome == "skipped"
    updated = service.upload(valid_product("OLD").product, "publish", DuplicateAction.UPDATE, conflict.existing_product)
    assert updated.outcome == "updated"
    assert client.updated[0][0] == 8
    assert "sku" not in client.updated[0][1]


def test_product_service_requires_a_decision_for_existing_sku() -> None:
    service = ProductUploadService(lambda: FakeClient({"OLD": {"id": 8}}))
    with pytest.raises(ValueError, match="decidir"):
        service.upload(valid_product("OLD").product, "draft")


def test_product_service_creates_a_product_without_sku() -> None:
    client = FakeClient()
    product = validate_product(
        ProductInput({"name": "Producto sin referencia", "regular_price": "10", "categories": "General"}),
        required_fields=("name", "regular_price", "categories"),
    ).product

    result = ProductUploadService(lambda: client).upload(product, "draft")

    assert result.outcome == "created"
    assert "sku" not in client.created[0]


def test_product_service_returns_remote_errors_as_a_typed_result() -> None:
    service = ProductUploadService(lambda: FakeClient(fail_skus={"FAIL"}))
    result = service.upload(valid_product("FAIL").product, "draft")
    assert result.outcome == "failed"
    assert result.error == "fallo remoto"


def test_batch_service_handles_conflicts_failures_and_cancellation() -> None:
    client = FakeClient({"UPDATE": {"id": 4}, "SKIP": {"id": 5}}, {"FAIL"})
    service = BatchUploadService(lambda: client)
    results = [valid_product("CREATE", 2), valid_product("UPDATE", 3), valid_product("SKIP", 4), valid_product("FAIL", 5)]
    conflicts = service.find_conflicts(results)
    assert set(conflicts) == {3, 4}
    result = service.upload(results, "draft", {3: DuplicateAction.UPDATE}, conflicts)
    assert result.counts == {"created": 1, "updated": 1, "skipped": 1, "failed": 1}
    assert "Fila 5: fallo remoto" in result.failures
    assert "sku" not in client.updated[0][1]
    stopped = service.upload(results, "draft", {}, conflicts, is_cancelled=lambda: True)
    assert stopped.cancelled is True
    assert stopped.rows == []


def test_batch_service_rejects_invalid_products_before_calling_client() -> None:
    client = FakeClient()
    invalid = validate_product(ProductInput({"sku": "X"}), required_fields=("name",))
    with pytest.raises(ValueError, match="no válidos"):
        BatchUploadService(lambda: client).upload([invalid], "draft", {}, {})
    assert client.created == []


def test_batch_service_safely_skips_a_duplicate_discovered_during_creation() -> None:
    class LateDuplicateClient(FakeClient):
        def __init__(self) -> None:
            super().__init__()
            self.create_attempted = False

        def find_by_sku(self, sku: str):
            return {"id": 9, "sku": sku} if self.create_attempted else None

        def create_product(self, payload, image_path=None):
            self.create_attempted = True
            raise WooCommerceError(
                "WooCommerce respondió 400: El producto con SKU ya está en la tabla de búsqueda",
                code="product_invalid_sku",
                status_code=400,
            )

    client = LateDuplicateClient()
    service = BatchUploadService(lambda: client)
    product = valid_product("LATE", 2)

    result = service.upload([product], "draft", {}, {})

    assert result.counts == {"created": 0, "updated": 0, "skipped": 1, "failed": 0}
    assert result.rows[0].detail == "SKU existente detectado durante la subida"


def test_batch_service_explains_an_orphaned_sku_lookup_entry() -> None:
    class OrphanSkuClient(FakeClient):
        def create_product(self, payload, image_path=None):
            raise WooCommerceError("El producto con SKU ya está en la tabla de búsqueda")

    service = BatchUploadService(lambda: OrphanSkuClient())

    result = service.upload([valid_product("ORPHAN", 2)], "draft", {}, {})

    assert result.counts["failed"] == 1
    assert "Revise la papelera" in result.rows[0].detail


def test_batch_service_sends_shield_location_taxonomies() -> None:
    client = FakeClient()
    shield = validate_product(ProductInput({
        "name": "Escudo de Sevilla",
        "sku": "ESC-1",
        "regular_price": "20",
        "categories": "Escudos",
        "ebdlt_pais": "españa",
        "ebdlt_region": "andalucía",
        "ebdlt_ciudad": "sevilla",
    }, row_number=2))

    result = BatchUploadService(lambda: client).upload([shield], "draft", {}, {})

    assert result.counts["created"] == 1
    assert client.created[0]["ebdlt_pais"] == ["España"]
    assert client.created[0]["ebdlt_region"] == ["Andalucía"]
    assert client.created[0]["ebdlt_ciudad"] == ["Sevilla"]


def test_batch_report_uses_historical_name_and_utf8_bom(tmp_path: Path) -> None:
    client = FakeClient({"SKIP": {"id": 2}})
    results = [valid_product("CREATE", 2), valid_product("SKIP", 3)]
    service = BatchUploadService(lambda: client)
    report = write_batch_report(service.upload(results, "draft", {}, service.find_conflicts(results)), tmp_path)
    assert report.name == "resultado_subida_lote.csv"
    data = report.read_bytes()
    assert data.startswith(b"\xef\xbb\xbf")
    assert "creado" in data.decode("utf-8-sig")
    assert "omitido" in data.decode("utf-8-sig")


def test_batch_reports_create_a_readable_excel_with_summary_and_detail(tmp_path: Path) -> None:
    client = FakeClient({"SKIP": {"id": 2}}, {"FAIL"})
    results = [valid_product("CREATE", 2), valid_product("SKIP", 3), valid_product("FAIL", 4)]
    service = BatchUploadService(lambda: client)
    batch = service.upload(results, "draft", {}, service.find_conflicts(results))

    reports = write_batch_reports(
        batch,
        csv_directory=tmp_path,
        xlsx_directory=tmp_path,
        timestamp=datetime(2026, 9, 2, 12, 30, 45),
    )

    assert reports.csv.name == "resultado_subida_lote.csv"
    assert reports.xlsx.name == "resultado_subida_lote_20260902_123045.xlsx"
    workbook = load_workbook(reports.xlsx, data_only=True)
    assert workbook.sheetnames == ["Resumen", "Detalle"]
    assert workbook["Resumen"]["B5"].value == 1
    assert workbook["Resumen"]["B7"].value == 1
    assert workbook["Resumen"]["B8"].value == 1
    assert list(workbook["Detalle"].values) == [
        ("Fila", "SKU", "Resultado", "Detalle"),
        (2, "CREATE", "Creado", None),
        (3, "SKIP", "Omitido", "SKU existente"),
        (4, "FAIL", "Error", "fallo remoto"),
    ]


def test_excel_report_uses_the_user_downloads_directory_by_default(tmp_path: Path, monkeypatch) -> None:
    downloads = tmp_path / "Descargas"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(report_service, "user_downloads_path", lambda: downloads)

    reports = write_batch_reports(
        BatchUploadService(lambda: FakeClient()).upload([valid_product("CREATE", 2)], "draft", {}, {}),
        timestamp=datetime(2026, 9, 2, 12, 30, 45),
    )

    assert reports.xlsx.parent == downloads
    assert reports.xlsx.exists()
